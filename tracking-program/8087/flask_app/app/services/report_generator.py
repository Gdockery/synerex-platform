"""
Reporting Engine™ — Phase 12

Spec: ECBS OS v4 §39, Figure A-14, Appendix B-24, C-26

Purpose
───────
Central report generation engine.  Assembles data from all ECBS modules
(Phases 7-11), formats it into structured payloads, and renders output files
in PDF, Excel, CSV, or JSON.

Seven Report Categories (spec §39)
────────────────────────────────────
1. Executive Summary™        — Annual savings, lifetime savings, capacity
                               recovered, DCV, ROI, payback, site health, alarms
2. Capacity & Performance™   — 5 capacity categories, utilization trends,
                               transformer analysis, loading profile
3. Power Quality™            — CBI score, harmonic/reactive/neutral burden,
                               THD trends, power factor history
4. Savings & Financials™     — 5 savings categories, waterfall, ROI/payback,
                               energy/demand/PF savings, CO₂ reduction
5. Environmental Impact™     — CO₂ reduction tons, carbon credit value,
                               sustainability metrics, Scope 2 impact
6. Alarms & Events™          — Active/resolved alarms by severity, MTTR,
                               top alarm types, alarm timeline, event log
7. Custom Report™            — User-selectable subset of the above

File outputs are saved under STORAGE_LOCAL_PATH/reports/{project_id}/
"""
from __future__ import annotations

import csv as csv_mod
import io
import json
import logging
import os
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from flask import current_app

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Storage helpers
# ─────────────────────────────────────────────────────────────────────────────

def _reports_dir(project_id: int) -> Path:
    storage = current_app.config.get("STORAGE_LOCAL_PATH", "/tmp")
    d = Path(storage) / "reports" / str(project_id)
    d.mkdir(parents=True, exist_ok=True)
    return d


def _file_path(project_id: int, report_id: int, fmt: str) -> Path:
    ext = {"pdf": "pdf", "excel": "xlsx", "csv": "csv", "json": "json"}.get(fmt, fmt)
    return _reports_dir(project_id) / f"report_{report_id}.{ext}"


def _file_url(project_id: int, report_id: int, fmt: str) -> str:
    ext = {"pdf": "pdf", "excel": "xlsx", "csv": "csv", "json": "json"}.get(fmt, fmt)
    return f"/files/reports/{project_id}/report_{report_id}.{ext}"


# ─────────────────────────────────────────────────────────────────────────────
# Data assembly — one function per category
# ─────────────────────────────────────────────────────────────────────────────

def _assembly_executive(project_id: int, from_ts: Optional[int],
                         to_ts: Optional[int]) -> dict:
    """Assemble Executive Summary™ data."""
    data = {
        "category":       "executive_summary",
        "annual_savings": None,
        "lifetime_savings": None,
        "capacity_recovered_kva": None,
        "deferred_capital_value": None,
        "roi": None,
        "payback": None,
        "site_health": None,
        "active_alarms": None,
        "cbi_score": None,
        "capacity_health_score": None,
    }
    try:
        from app.models.savings_intelligence import SavingsIntelligence
        si = (SavingsIntelligence.query
              .filter_by(project_id=project_id, isDeleted=False)
              .order_by(SavingsIntelligence.bucket_ts.desc())
              .first())
        if si:
            data["annual_savings"]   = si.annual_savings
            data["lifetime_savings"] = si.lifetime_savings
            data["roi"]              = si.roi
            data["payback"]          = si.payback
    except Exception as exc:
        logger.debug("exec: savings error: %s", exc)

    try:
        from app.models.capacity_intelligence import CapacityIntelligence
        ci = (CapacityIntelligence.query
              .filter_by(project_id=project_id, isDeleted=False)
              .order_by(CapacityIntelligence.bucket_ts.desc())
              .first())
        if ci:
            data["capacity_recovered_kva"]  = ci.recoverable_capacity
            data["deferred_capital_value"]  = ci.deferred_capital_value
            data["capacity_health_score"]   = ci.capacity_health_score
    except Exception as exc:
        logger.debug("exec: capacity error: %s", exc)

    try:
        from app.models.current_balance_metrics import CurrentBalanceMetrics
        cbi = (CurrentBalanceMetrics.query
               .filter_by(project_id=project_id, isDeleted=False)
               .order_by(CurrentBalanceMetrics.bucket_ts.desc())
               .first())
        if cbi:
            data["cbi_score"]   = getattr(cbi, "cbi_score", None)
            data["site_health"] = data["cbi_score"]
    except Exception as exc:
        logger.debug("exec: cbi error: %s", exc)

    try:
        from app.models.alarm import Alarm
        data["active_alarms"] = Alarm.query.filter(
            Alarm.project_id == project_id,
            Alarm.status.notin_(["resolved", "closed"]),
            Alarm.isDeleted == False,
        ).count()
    except Exception as exc:
        logger.debug("exec: alarms error: %s", exc)

    return data


def _assembly_capacity(project_id: int, from_ts: Optional[int],
                        to_ts: Optional[int]) -> dict:
    """Assemble Capacity & Performance™ data."""
    data: dict = {
        "category": "capacity_performance",
        "summary": {},
        "trends": [],
    }
    try:
        from app.services.capacity_intelligence_engine import dashboard_summary
        data["summary"] = dashboard_summary(project_id)
    except Exception as exc:
        logger.debug("capacity: summary error: %s", exc)

    try:
        from app.models.capacity_intelligence import CapacityIntelligence
        q = (CapacityIntelligence.query
             .filter_by(project_id=project_id, isDeleted=False)
             .order_by(CapacityIntelligence.bucket_ts.desc())
             .limit(96))   # last 24h at 15-min intervals
        if from_ts:
            q = q.filter(CapacityIntelligence.bucket_ts >= from_ts)
        if to_ts:
            q = q.filter(CapacityIntelligence.bucket_ts <= to_ts)
        rows = q.all()
        data["trends"] = [
            {
                "ts":                r.bucket_ts,
                "utilization_pct":   r.utilization_pct,
                "hidden_pct":        r.hidden_pct,
                "recoverable_pct":   r.recoverable_pct,
                "health_score":      r.capacity_health_score,
                "available_kva":     r.available_capacity,
                "recoverable_kva":   r.recoverable_capacity,
            }
            for r in rows
        ]
    except Exception as exc:
        logger.debug("capacity: trends error: %s", exc)

    return data


def _assembly_power_quality(project_id: int, from_ts: Optional[int],
                             to_ts: Optional[int]) -> dict:
    """Assemble Power Quality™ data."""
    data: dict = {
        "category": "power_quality",
        "summary": {},
        "trends": [],
    }
    try:
        from app.models.current_balance_metrics import CurrentBalanceMetrics
        q = (CurrentBalanceMetrics.query
             .filter_by(project_id=project_id, isDeleted=False)
             .order_by(CurrentBalanceMetrics.bucket_ts.desc())
             .limit(96))
        if from_ts:
            q = q.filter(CurrentBalanceMetrics.bucket_ts >= from_ts)
        if to_ts:
            q = q.filter(CurrentBalanceMetrics.bucket_ts <= to_ts)
        rows = q.all()

        if rows:
            latest = rows[0]
            data["summary"] = {
                "cbi_score":           getattr(latest, "cbi_score", None),
                "harmonic_burden_pct": getattr(latest, "harmonic_burden_pct", None),
                "reactive_burden_pct": getattr(latest, "reactive_burden_pct", None),
                "neutral_burden_pct":  getattr(latest, "neutral_burden_pct", None),
                "avg_pf":              getattr(latest, "avg_pf", None),
            }
        data["trends"] = [
            {
                "ts":                  r.bucket_ts,
                "cbi_score":           getattr(r, "cbi_score", None),
                "harmonic_burden_pct": getattr(r, "harmonic_burden_pct", None),
                "reactive_burden_pct": getattr(r, "reactive_burden_pct", None),
                "neutral_burden_pct":  getattr(r, "neutral_burden_pct", None),
                "avg_pf":              getattr(r, "avg_pf", None),
                "avg_kw":              getattr(r, "avg_kw", None),
                "avg_kva":             getattr(r, "avg_kva", None),
            }
            for r in rows
        ]
    except Exception as exc:
        logger.debug("pq: error: %s", exc)

    return data


def _assembly_savings(project_id: int, from_ts: Optional[int],
                       to_ts: Optional[int]) -> dict:
    """Assemble Savings & Financials™ data."""
    data: dict = {
        "category": "savings_financials",
        "summary": {},
        "trends": [],
        "waterfall": [],
    }
    try:
        from app.services.savings_intelligence_engine import dashboard_summary
        data["summary"] = dashboard_summary(project_id)
    except Exception as exc:
        logger.debug("savings: summary error: %s", exc)

    try:
        from app.models.savings_intelligence import SavingsIntelligence
        q = (SavingsIntelligence.query
             .filter_by(project_id=project_id, isDeleted=False)
             .order_by(SavingsIntelligence.bucket_ts.desc())
             .limit(96))
        if from_ts:
            q = q.filter(SavingsIntelligence.bucket_ts >= from_ts)
        if to_ts:
            q = q.filter(SavingsIntelligence.bucket_ts <= to_ts)
        rows = q.all()
        data["trends"] = [
            {
                "ts":            r.bucket_ts,
                "annual_savings": r.annual_savings,
                "roi":           r.roi,
                "payback":       r.payback,
                "energy_savings": r.energy_savings,
                "demand_savings": r.demand_savings,
                "pf_savings":    r.pf_savings,
                "capacity_value": r.capacity_value,
                "sustainability_value": r.sustainability_value,
            }
            for r in rows
        ]
        if rows:
            latest = rows[0]
            data["waterfall"] = [
                {"label": "Energy Savings",       "value": latest.energy_savings},
                {"label": "Demand Savings",        "value": latest.demand_savings},
                {"label": "Power Factor Savings",  "value": latest.pf_savings},
                {"label": "Capacity Value",        "value": latest.capacity_value},
                {"label": "Sustainability Value",  "value": latest.sustainability_value},
                {"label": "Total Annual Savings",  "value": latest.annual_savings},
            ]
    except Exception as exc:
        logger.debug("savings: trends error: %s", exc)

    return data


def _assembly_environmental(project_id: int, from_ts: Optional[int],
                             to_ts: Optional[int]) -> dict:
    """Assemble Environmental Impact™ data."""
    data: dict = {
        "category": "environmental_impact",
        "co2_reduction_tons": None,
        "carbon_credit_value": None,
        "kwh_reduction_per_year": None,
        "recoverable_kva": None,
        "trends": [],
    }
    try:
        from app.models.savings_intelligence import SavingsIntelligence
        latest = (SavingsIntelligence.query
                  .filter_by(project_id=project_id, isDeleted=False)
                  .order_by(SavingsIntelligence.bucket_ts.desc())
                  .first())
        if latest:
            data["co2_reduction_tons"]    = latest.co2_reduction_tons
            data["kwh_reduction_per_year"] = latest.kwh_per_year
            data["recoverable_kva"]       = latest.recoverable_kva
            carbon_price = getattr(latest, "carbon_credit_price", None) or 25.0
            if latest.co2_reduction_tons:
                data["carbon_credit_value"] = round(latest.co2_reduction_tons * carbon_price, 2)
    except Exception as exc:
        logger.debug("env: error: %s", exc)

    return data


def _assembly_alarms(project_id: int, from_ts: Optional[int],
                     to_ts: Optional[int]) -> dict:
    """Assemble Alarms & Events™ data."""
    data: dict = {
        "category": "alarms_events",
        "summary": {},
        "active_alarms": [],
        "recent_events": [],
    }
    try:
        from app.services.alarm_engine import dashboard_summary
        data["summary"] = dashboard_summary(project_id)
    except Exception as exc:
        logger.debug("alarms: summary error: %s", exc)

    try:
        from app.models.alarm import Alarm
        q = (Alarm.query
             .filter_by(project_id=project_id, isDeleted=False)
             .filter(Alarm.status.notin_(["resolved", "closed"]))
             .order_by(Alarm.triggered_at.desc())
             .limit(50))
        data["active_alarms"] = [
            {
                "id":          a.id,
                "alarm_type":  a.alarm_type,
                "severity":    a.severity,
                "title":       a.title,
                "status":      a.status,
                "triggered_at": a.triggered_at,
            }
            for a in q.all()
        ]
    except Exception as exc:
        logger.debug("alarms: active error: %s", exc)

    try:
        from app.models.alarm import Event
        q = (Event.query
             .filter_by(project_id=project_id)
             .order_by(Event.event_ts.desc())
             .limit(100))
        if from_ts:
            q = q.filter(Event.event_ts >= from_ts)
        data["recent_events"] = [
            {
                "source":     ev.source,
                "event_type": ev.event_type,
                "title":      ev.title,
                "severity":   ev.severity,
                "event_ts":   ev.event_ts,
            }
            for ev in q.all()
        ]
    except Exception as exc:
        logger.debug("alarms: events error: %s", exc)

    return data


def _assemble_report_data(category: str, project_id: int,
                           from_ts: Optional[int], to_ts: Optional[int],
                           custom_sections: Optional[list] = None) -> dict:
    """Route to the correct assembly function."""
    assemblers = {
        "executive_summary":   _assembly_executive,
        "capacity_performance": _assembly_capacity,
        "power_quality":       _assembly_power_quality,
        "savings_financials":  _assembly_savings,
        "environmental_impact": _assembly_environmental,
        "alarms_events":       _assembly_alarms,
    }

    if category == "custom" and custom_sections:
        combined = {"category": "custom", "sections": {}}
        for sec in custom_sections:
            fn = assemblers.get(sec)
            if fn:
                combined["sections"][sec] = fn(project_id, from_ts, to_ts)
        return combined

    fn = assemblers.get(category, _assembly_executive)
    return fn(project_id, from_ts, to_ts)


# ─────────────────────────────────────────────────────────────────────────────
# Format renderers
# ─────────────────────────────────────────────────────────────────────────────

def _render_json(data: dict, file_path: Path) -> int:
    """Write JSON report. Returns file size in bytes."""
    content = json.dumps(data, indent=2, default=str)
    file_path.write_text(content, encoding="utf-8")
    return file_path.stat().st_size


def _render_csv(data: dict, file_path: Path) -> int:
    """Write CSV report from trend rows (first list found in data)."""
    rows = None
    for v in data.values():
        if isinstance(v, list) and v and isinstance(v[0], dict):
            rows = v
            break

    buf = io.StringIO()
    if rows:
        writer = csv_mod.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    else:
        # No list data — write summary as key-value CSV
        writer = csv_mod.writer(buf)
        writer.writerow(["Field", "Value"])
        for k, v in data.items():
            if not isinstance(v, (list, dict)):
                writer.writerow([k, v])

    file_path.write_text(buf.getvalue(), encoding="utf-8")
    return file_path.stat().st_size


def _render_excel(data: dict, file_path: Path, report_name: str) -> int:
    """Write Excel report using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed — falling back to CSV for Excel export")
        csv_path = file_path.with_suffix(".csv")
        _render_csv(data, csv_path)
        csv_path.rename(file_path)
        return file_path.stat().st_size

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1a5276")
    center_align = Alignment(horizontal="center")

    # Title row
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value     = report_name
    title_cell.font      = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor="0e3460")
    title_cell.alignment = center_align

    row = 3

    def _write_section_header(label: str):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=12)
        row += 1

    def _write_kv(key: str, val):
        nonlocal row
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=val)
        row += 1

    def _write_table(rows_list: list):
        nonlocal row
        if not rows_list or not isinstance(rows_list[0], dict):
            return
        headers = list(rows_list[0].keys())
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_i, value=h)
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = center_align
        row += 1
        for r in rows_list:
            for col_i, h in enumerate(headers, 1):
                ws.cell(row=row, column=col_i, value=r.get(h))
            row += 1
        row += 1

    # Write summary keys
    summary = data.get("summary", {})
    if summary:
        _write_section_header("Summary")
        for k, v in summary.items():
            if not isinstance(v, (list, dict)):
                _write_kv(k, v)
        row += 1

    # Write simple scalar fields from root
    scalars = {k: v for k, v in data.items()
               if k not in ("summary", "trends", "waterfall", "category", "sections")
               and not isinstance(v, (list, dict))}
    if scalars:
        _write_section_header("Metrics")
        for k, v in scalars.items():
            _write_kv(k, v)
        row += 1

    # Write trends table
    trends = data.get("trends") or data.get("waterfall")
    if trends:
        _write_section_header("Data")
        _write_table(trends)

    # Custom report: write each section on its own sheet
    sections = data.get("sections", {})
    for sec_name, sec_data in sections.items():
        sheet = wb.create_sheet(title=sec_name[:31])
        sheet.cell(row=1, column=1, value=sec_name).font = Font(bold=True)
        sec_row = 3
        for k, v in sec_data.items():
            if isinstance(v, list) and v:
                sheet.cell(row=sec_row, column=1, value=k).font = Font(bold=True)
                sec_row += 1
                if isinstance(v[0], dict):
                    hdrs = list(v[0].keys())
                    for ci, h in enumerate(hdrs, 1):
                        sheet.cell(row=sec_row, column=ci, value=h).font = header_font
                    sec_row += 1
                    for r in v:
                        for ci, h in enumerate(hdrs, 1):
                            sheet.cell(row=sec_row, column=ci, value=r.get(h))
                        sec_row += 1
            elif not isinstance(v, dict):
                sheet.cell(row=sec_row, column=1, value=k)
                sheet.cell(row=sec_row, column=2, value=v)
                sec_row += 1

    # Auto-size columns
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value), default=10
            )
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(str(file_path))
    return file_path.stat().st_size


def _render_pdf(data: dict, file_path: Path, report_name: str,
                project_id: int) -> int:
    """
    Render PDF report.

    Builds a simple HTML page from the assembled data and renders it with
    WeasyPrint (already available in requirements.txt).
    Falls back to JSON if WeasyPrint is unavailable.
    """
    try:
        import weasyprint
    except ImportError:
        logger.warning("weasyprint unavailable — falling back to JSON for PDF export")
        json_path = file_path.with_suffix(".json")
        _render_json(data, json_path)
        json_path.rename(file_path)
        return file_path.stat().st_size

    ts_label = datetime.now(timezone.utc).strftime("%B %d, %Y %H:%M UTC")
    category_labels = {
        "executive_summary":    "Executive Summary™",
        "capacity_performance": "Capacity & Performance™",
        "power_quality":        "Power Quality™",
        "savings_financials":   "Savings & Financials™",
        "environmental_impact": "Environmental Impact™",
        "alarms_events":        "Alarms & Events™",
        "custom":               "Custom Report™",
    }
    cat_label = category_labels.get(data.get("category", ""), "ECBS Report")

    def _fmt(val, decimals=2, suffix=""):
        if val is None:
            return "—"
        try:
            return f"{float(val):,.{decimals}f}{suffix}"
        except Exception:
            return str(val)

    # Build metric cards HTML
    card_html = ""
    summary = data.get("summary", {})
    scalars  = {k: v for k, v in data.items()
                if k not in ("summary", "trends", "waterfall", "category", "sections")
                and not isinstance(v, (list, dict))}
    all_metrics = {**scalars, **summary}

    for k, v in all_metrics.items():
        if isinstance(v, (list, dict)):
            continue
        label = k.replace("_", " ").title()
        val   = _fmt(v)
        card_html += f'<div class="card"><div class="label">{label}</div><div class="value">{val}</div></div>\n'

    # Waterfall / trends table
    table_html = ""
    rows_data = data.get("waterfall") or data.get("trends") or data.get("active_alarms") or []
    if rows_data and isinstance(rows_data[0], dict):
        headers = list(rows_data[0].keys())
        table_html = "<table><tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"
        for r in rows_data[:100]:
            table_html += "<tr>" + "".join(f"<td>{_fmt(r.get(h)) if isinstance(r.get(h), (int, float)) else (r.get(h) or '—')}</td>" for h in headers) + "</tr>"
        table_html += "</table>"

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; color: #1a1a2e; margin: 0; padding: 20px; }}
  .header {{ background: #0e3460; color: white; padding: 24px 32px; margin: -20px -20px 24px; }}
  .header h1 {{ margin: 0 0 4px; font-size: 22px; }}
  .header .sub {{ font-size: 13px; opacity: 0.8; }}
  .cards {{ display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 24px; }}
  .card {{ background: #f4f6fb; border-left: 4px solid #1abc9c; padding: 12px 16px;
           border-radius: 4px; min-width: 160px; flex: 1; }}
  .card .label {{ font-size: 11px; color: #7f8c8d; text-transform: uppercase; margin-bottom: 4px; }}
  .card .value {{ font-size: 20px; font-weight: bold; color: #0e3460; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  th {{ background: #1a5276; color: white; padding: 8px; text-align: left; }}
  td {{ padding: 7px 8px; border-bottom: 1px solid #e8ecf0; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
  .footer {{ margin-top: 32px; font-size: 11px; color: #95a5a6; text-align: center; }}
</style>
</head>
<body>
  <div class="header">
    <h1>SYNEREX Energy Intelligence Portal</h1>
    <div class="sub">{cat_label} &nbsp;|&nbsp; {report_name} &nbsp;|&nbsp; Generated {ts_label}</div>
  </div>
  <div class="cards">
    {card_html}
  </div>
  {table_html}
  <div class="footer">ECBS Intelligence Platform™ &nbsp;•&nbsp; Confidential &nbsp;•&nbsp; {ts_label}</div>
</body>
</html>"""

    weasyprint.HTML(string=html).write_pdf(str(file_path))
    return file_path.stat().st_size


# ─────────────────────────────────────────────────────────────────────────────
# Main generation entry point
# ─────────────────────────────────────────────────────────────────────────────

def generate_report(report_id: int) -> dict:
    """
    Generate a report for the given EcbsReport.id.

    Loads the EcbsReport record, assembles data, renders the requested format,
    saves the file, creates a ReportExport row, and marks the report complete.

    Returns a summary dict.
    """
    from app.extensions import db
    from app.models.ecbs_report import (
        EcbsReport, ReportExport,
        STATUS_RUNNING, STATUS_COMPLETE, STATUS_FAILED,
    )

    report = EcbsReport.query.get(report_id)
    if not report:
        raise ValueError(f"EcbsReport id={report_id} not found")

    now_ms = int(time.time() * 1000)
    report.status    = STATUS_RUNNING
    report.updatedAt = now_ms
    db.session.commit()

    try:
        # 1. Assemble data
        data = _assemble_report_data(
            category=report.category,
            project_id=report.project_id,
            from_ts=report.from_date,
            to_ts=report.to_date,
        )

        # 2. Render to file
        fmt       = report.format
        fpath     = _file_path(report.project_id, report.id, fmt)
        name      = report.name

        if fmt == "json":
            size = _render_json(data, fpath)
        elif fmt == "csv":
            size = _render_csv(data, fpath)
        elif fmt == "excel":
            size = _render_excel(data, fpath, name)
        else:  # pdf (default)
            size = _render_pdf(data, fpath, name, report.project_id)

        # 3. Create ReportExport record
        export = ReportExport(
            report_id=report.id,
            format=fmt,
            file_path=str(fpath),
            file_url=_file_url(report.project_id, report.id, fmt),
            file_size=size,
            download_count=0,
            created_at=now_ms,
            createdAt=now_ms,
            updatedAt=now_ms,
        )
        db.session.add(export)

        # 4. Mark report complete
        report.status       = STATUS_COMPLETE
        report.generated_at = now_ms
        report.updatedAt    = now_ms
        db.session.commit()

        logger.info("report generated id=%d format=%s size=%d", report.id, fmt, size)
        return {"id": report.id, "status": STATUS_COMPLETE, "file_url": export.file_url, "size": size}

    except Exception as exc:
        logger.error("report generation error id=%d: %s", report_id, exc)
        report.status    = STATUS_FAILED
        report.error     = str(exc)
        report.updatedAt = int(time.time() * 1000)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
        raise


def create_and_generate(
    project_id: int,
    category: str,
    fmt: str,
    name: Optional[str] = None,
    from_ts: Optional[int] = None,
    to_ts:   Optional[int] = None,
    site_id: Optional[int] = None,
    generated_by: Optional[int] = None,
    schedule_id:  Optional[int] = None,
) -> dict:
    """
    Convenience: create an EcbsReport record and immediately generate it.
    Used by both the API route and the scheduled runner.
    """
    from app.extensions import db
    from app.models.ecbs_report import EcbsReport, STATUS_PENDING

    now_ms   = int(time.time() * 1000)
    cat_labels = {
        "executive_summary":    "Executive Summary",
        "capacity_performance": "Capacity & Performance",
        "power_quality":        "Power Quality",
        "savings_financials":   "Savings & Financials",
        "environmental_impact": "Environmental Impact",
        "alarms_events":        "Alarms & Events",
        "custom":               "Custom Report",
    }
    default_name = cat_labels.get(category, category.replace("_", " ").title())
    ts_suffix    = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    report = EcbsReport(
        project_id=project_id,
        site_id=site_id,
        name=name or f"{default_name} — {ts_suffix}",
        category=category,
        report_type=category,
        format=fmt,
        status=STATUS_PENDING,
        from_date=from_ts,
        to_date=to_ts,
        generated_by=generated_by,
        schedule_id=schedule_id,
        createdAt=now_ms,
        updatedAt=now_ms,
    )
    db.session.add(report)
    db.session.commit()

    return generate_report(report.id)
